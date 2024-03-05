from collections import OrderedDict
from pathlib import Path

import openpyxl


class XLSXFile:
    """ Excel-файл, содержащий 1 таблицу с заголовком """

    def __init__(self, path: Path, auto_create=True) -> None:
        if path.suffix != ".xlsx":
            raise Exception("The file extension must be .xlsx")
        self._path = path
        if auto_create and not self.exists():
            self.write([])

    @property
    def path(self) -> Path:
        return self._path

    def exists(self) -> bool:
        return self._path.exists()

    def read(self) -> list[OrderedDict]:

        wb = openpyxl.load_workbook(str(self._path))
        ws = wb.active
        if ws.max_row == 0:
            return []

        keys = []
        for column in range(1, ws.max_column + 1):
            keys.append(ws.cell(1, column).value)

        data = []
        for row in range(2, ws.max_row + 1):
            dct = OrderedDict()
            for column, key in enumerate(keys, start=1):
                dct[key] = ws.cell(row, column).value
            data.append(dct)

        return data

    def write(self, data: list[OrderedDict]) -> None:

        wb = openpyxl.Workbook()
        ws = wb.active
        if not data:
            wb.save(str(self._path))
            return

        for column, key in enumerate(data[0].keys(), start=1):
            ws.cell(1, column, key)

        for row, dct in enumerate(data, start=2):
            for column, value in enumerate(dct.values(), start=1):
                ws.cell(row, column, value)

        self._adjust_columns(ws)
        wb.save(str(self._path))

    @staticmethod
    def _adjust_columns(sheet) -> None:
        """
        Отрегулировать ширину столбцов листа
        https://stackoverflow.com/questions/39529662/python-automatically-adjust-width-of-an-excel-files-columns
        """
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length
            sheet.column_dimensions[column].width = adjusted_width
